"""
Export supertrend scanner CSV to Excel with two sheets:
  Sheet 1 (Data Dictionary): Field name + description for every column
  Sheet 2 (Touch Data):      The actual touch-event data

Usage:  python3 export_excel.py [csv_path] [output_xlsx]
"""

import sys

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

# ── Column descriptions ──────────────────────────────────────────────

FIELD_DESCRIPTIONS = {
    # Core OHLCV
    "timestamp": "Bar timestamp (exchange timezone)",
    "symbol": "Instrument symbol (NIFTY, BANKNIFTY, etc.)",
    "open": "Bar open price",
    "high": "Bar high price",
    "low": "Bar low price",
    "close": "Bar close price",
    "volume": "Bar traded volume",
    "oi": "Open interest (futures only, 0 for index)",
    "oi_change": "Change in open interest from previous bar",
    # VIX
    "india_vix": "India VIX (daily, forward-filled onto intraday bars)",
    # Supertrend
    "st_upperband": "Supertrend upper band value",
    "st_lowerband": "Supertrend lower band value",
    "st_trend": "Supertrend trend direction: 1 = uptrend (green), -1 = downtrend (red)",
    "st_line": "Active supertrend line (lower band in uptrend, upper in downtrend)",
    # Momentum / Trend indicators
    "rsi": "Relative Strength Index (14-period)",
    "rsi_sma": "SMA of Relative Strength Index (14-period)",
    "atr": "Average True Range (14-period)",
    "macd": "MACD line (12/26 EMA difference)",
    "macd_signal": "MACD signal line (9-period EMA of MACD)",
    "macd_hist": "MACD histogram (MACD - signal)",
    "adx": "Average Directional Index (trend strength, 14-period)",
    "plus_di": "+DI (positive directional indicator)",
    "minus_di": "-DI (negative directional indicator)",
    "cci": "Commodity Channel Index (20-period)",
    # Moving averages
    "ema_5": "Exponential Moving Average (5-period)",
    "ema_9": "Exponential Moving Average (9-period)",
    "ema_13": "Exponential Moving Average (13-period)",
    "ema_21": "Exponential Moving Average (21-period)",
    "ema_34": "Exponential Moving Average (34-period)",
    "ema_50": "Exponential Moving Average (50-period)",
    "ema_100": "Exponential Moving Average (100-period)",
    "ema_200": "Exponential Moving Average (200-period)",
    "sma_20": "Simple Moving Average (20-period)",
    "sma_50": "Simple Moving Average (50-period)",
    "sma_100": "Simple Moving Average (100-period)",
    "sma_200": "Simple Moving Average (200-period)",
    # Bollinger Bands
    "bb_mid": "Bollinger Band middle (20-period SMA)",
    "bb_upper": "Bollinger Band upper (mid + 2 std dev)",
    "bb_lower": "Bollinger Band lower (mid - 2 std dev)",
    "bb_pctb": "Bollinger %B: (close - lower) / (upper - lower). >1 = above upper, <0 = below lower",
    "bb_bandwidth": "Bollinger Bandwidth: (upper - lower) / mid. Squeeze when low",
    # Stochastic
    "stoch_k": "Stochastic %K (14-period, smoothed)",
    "stoch_d": "Stochastic %D (3-period SMA of %K)",
    # Volume
    "vwap": "Session VWAP (resets daily)",
    "vol_sma": "Volume SMA (20-period average volume)",
    "vol_ratio": "Volume ratio: current volume / vol_sma. >1 = above average volume",
    "obv": "On-Balance Volume (cumulative volume * price direction)",
    # Candlestick patterns
    "candle_pattern": "Detected candlestick pattern name (BULLISH_MARUBOZU, HAMMER, DOJI, ENGULFING, etc.)",
    "candle_signal": "Pattern signal: BULLISH, BEARISH, or NEUTRAL",
    "candle_body_pct": "Candle body as % of total range (0-100). Small = indecision, large = conviction",
    # Divergences
    "div_rsi_bull": "RSI bullish divergence detected (price lower low, RSI higher low)",
    "div_rsi_bear": "RSI bearish divergence detected (price higher high, RSI lower high)",
    "div_macd_bull": "MACD histogram bullish divergence detected",
    "div_macd_bear": "MACD histogram bearish divergence detected",
    "div_stoch_bull": "Stochastic %K bullish divergence detected",
    "div_stoch_bear": "Stochastic %K bearish divergence detected",
    "div_adx_bull": "ADX bullish divergence detected",
    "div_adx_bear": "ADX bearish divergence detected",
    "div_cci_bull": "CCI bullish divergence detected",
    "div_cci_bear": "CCI bearish divergence detected",
    "div_obv_bull": "OBV bullish divergence detected",
    "div_obv_bear": "OBV bearish divergence detected",
    "divergence_summary": "Comma-separated list of active divergences (e.g. RSI_BULL,MACD_BEAR)",
    # Fibonacci
    "fib_swing_high": "Swing high of current supertrend trend segment (Fib anchor)",
    "fib_swing_low": "Swing low of current supertrend trend segment (Fib anchor)",
    "fib_0.0": "Fibonacci 0% level (swing extreme = trend start)",
    "fib_0.236": "Fibonacci 23.6% retracement level",
    "fib_0.382": "Fibonacci 38.2% retracement level",
    "fib_0.5": "Fibonacci 50% retracement level",
    "fib_0.618": "Fibonacci 61.8% retracement level (golden ratio)",
    "fib_0.786": "Fibonacci 78.6% retracement level",
    "fib_1.0": "Fibonacci 100% retracement level (full retracement)",
    "fib_1.272": "Fibonacci 127.2% extension level",
    "fib_1.618": "Fibonacci 161.8% extension level",
    "fib_2.0": "Fibonacci 200% extension level",
    "fib_2.618": "Fibonacci 261.8% extension level",
    "fib_nearest_level": "Which Fib ratio the current close is nearest to (e.g. 0.618)",
    "fib_distance_pct": "% distance from the nearest Fibonacci level",
    # Touch counting
    "trend_touch_number": "Which touch # this is within the current trend segment (1st, 2nd, 3rd...)",
    "trend_start_idx": "Positional index where the current supertrend trend segment began",
    "bars_since_trend_start": "Number of bars elapsed since the current trend started",
    # Touch detection
    "active_band": "The supertrend band value being tested (lower in uptrend, upper in downtrend)",
    "band_distance_pct": "% distance between close and the active supertrend band",
    "signal_type": "Touch signal: BUY_TOUCH (uptrend touch) or SELL_TOUCH (downtrend touch)",
    # Touch outcomes
    "bounced": "Did price bounce the threshold before trend reversed? True/False",
    "bounce_pct": "Peak bounce as % of the touch price",
    "peak_bounce_points": "Best favorable move (in points) reached before the trend reversed",
    "peak_bounce_timestamp": "Timestamp when the peak bounce occurred",
    "points_at_reversal": "Points move from touch close to reversal close (favorable = positive)",
    "reversal_timestamp": "Timestamp when the supertrend trend actually flipped",
    "bars_to_reversal": "Number of bars from touch to the trend reversal (NaN if no reversal)",
    "bounce_threshold_points": "The bounce threshold used for this symbol (NIFTY=50, BANKNIFTY=300 pts)",
    "bounce_threshold_hit": "Same as 'bounced': did peak bounce reach the threshold before reversal?",
    "bars_to_threshold_hit": "How many bars it took to reach the bounce threshold (NaN if never)",
    # Fractal Zones
    "nearest_fractal_zone_price": "Price of the nearest valid fractal S/R zone (wick edge)",
    "nearest_fractal_zone_type": "Type of the nearest zone (Resistance / Support)",
    "fractal_zone_distance_pct": "% distance from touch's active_band to the zone",
    "fractal_zone_strength_score": "Composite strength score: rejection * (1/width) * (1 + 0.1*retests)",
    "fractal_zone_strength_bucket": "Quartile bucket of the strength score (weak, medium, strong, very_strong)",
}


def export_to_excel(csv_path: str, xlsx_path: str):
    """Read CSV and export to Excel with data dictionary + data sheets."""
    print(f"Reading {csv_path} ...")
    df = pd.read_csv(csv_path)
    print(f"  {len(df)} rows, {len(df.columns)} columns")

    wb = Workbook()

    # ── Sheet 1: Data Dictionary ──────────────────────────────────────

    ws_dict = wb.active
    ws_dict.title = "Data Dictionary"

    # Styles
    header_font = Font(name="Calibri", bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    cat_font = Font(name="Calibri", bold=True, size=11, color="2F5496")
    cat_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    thin_border = Border(
        bottom=Side(style="thin", color="B4C6E7"),
    )

    # Headers
    ws_dict.append(["#", "Field Name", "Description", "Sample Value"])
    for col_idx in range(1, 5):
        cell = ws_dict.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="left")

    # Column widths
    ws_dict.column_dimensions["A"].width = 5
    ws_dict.column_dimensions["B"].width = 28
    ws_dict.column_dimensions["C"].width = 80
    ws_dict.column_dimensions["D"].width = 30

    # Categories for grouping
    categories = {
        "OHLCV & Market Data": [
            "timestamp",
            "symbol",
            "open",
            "high",
            "low",
            "close",
            "volume",
            "oi",
            "oi_change",
            "india_vix",
        ],
        "Supertrend": ["st_upperband", "st_lowerband", "st_trend", "st_line"],
        "Momentum & Trend": [
            "rsi",
            "rsi_sma",
            "atr",
            "macd",
            "macd_signal",
            "macd_hist",
            "adx",
            "plus_di",
            "minus_di",
            "cci",
        ],
        "Moving Averages": [
            "ema_5",
            "ema_9",
            "ema_13",
            "ema_21",
            "ema_34",
            "ema_50",
            "ema_100",
            "ema_200",
            "sma_20",
            "sma_50",
            "sma_100",
            "sma_200",
        ],
        "Bollinger Bands": ["bb_mid", "bb_upper", "bb_lower", "bb_pctb", "bb_bandwidth"],
        "Stochastic": ["stoch_k", "stoch_d"],
        "Volume": ["vwap", "vol_sma", "vol_ratio", "obv"],
        "Candlestick Patterns": ["candle_pattern", "candle_signal", "candle_body_pct"],
        "Divergences": [
            "div_rsi_bull",
            "div_rsi_bear",
            "div_macd_bull",
            "div_macd_bear",
            "div_stoch_bull",
            "div_stoch_bear",
            "div_adx_bull",
            "div_adx_bear",
            "div_cci_bull",
            "div_cci_bear",
            "div_obv_bull",
            "div_obv_bear",
            "divergence_summary",
        ],
        "Fibonacci Levels": [
            "fib_swing_high",
            "fib_swing_low",
            "fib_0.0",
            "fib_0.236",
            "fib_0.382",
            "fib_0.5",
            "fib_0.618",
            "fib_0.786",
            "fib_1.0",
            "fib_1.272",
            "fib_1.618",
            "fib_2.0",
            "fib_2.618",
            "fib_nearest_level",
            "fib_distance_pct",
        ],
        "Fractal Zones": [
            "nearest_fractal_zone_price",
            "nearest_fractal_zone_type",
            "fractal_zone_distance_pct",
            "fractal_zone_strength_score",
            "fractal_zone_strength_bucket",
        ],
        "Touch Counting": ["trend_touch_number", "trend_start_idx", "bars_since_trend_start"],
        "Touch Detection": ["active_band", "band_distance_pct", "signal_type"],
        "Touch Outcomes": [
            "bounced",
            "bounce_pct",
            "peak_bounce_points",
            "peak_bounce_timestamp",
            "points_at_reversal",
            "reversal_timestamp",
            "bars_to_reversal",
            "bounce_threshold_points",
            "bounce_threshold_hit",
            "bars_to_threshold_hit",
        ],
    }

    row_num = 1
    field_num = 0
    for category, fields in categories.items():
        row_num += 1
        # Category header row
        ws_dict.append([None, f"── {category} ──", None, None])
        for col_idx in range(1, 5):
            cell = ws_dict.cell(row=row_num, column=col_idx)
            cell.font = cat_font
            cell.fill = cat_fill

        for field in fields:
            if field not in df.columns:
                continue
            row_num += 1
            field_num += 1
            desc = FIELD_DESCRIPTIONS.get(field, "")
            # Get a non-null sample value
            sample = df[field].dropna()
            sample_val = str(sample.iloc[0]) if len(sample) > 0 else ""
            if len(sample_val) > 40:
                sample_val = sample_val[:40] + "..."

            ws_dict.append([field_num, field, desc, sample_val])
            for col_idx in range(1, 5):
                ws_dict.cell(row=row_num, column=col_idx).border = thin_border

    # Any columns not categorized above
    categorized = set()
    for fields in categories.values():
        categorized.update(fields)

    uncategorized = [c for c in df.columns if c not in categorized]

    # Handle HTF columns separately
    htf_cols = [c for c in uncategorized if c.startswith("htf_")]
    if htf_cols:
        row_num += 1
        ws_dict.append([None, "── HTF Context ──", None, None])
        for col_idx in range(1, 5):
            cell = ws_dict.cell(row=row_num, column=col_idx)
            cell.font = cat_font
            cell.fill = cat_fill
        for field in htf_cols:
            row_num += 1
            field_num += 1
            sample = df[field].dropna()
            sample_val = str(sample.iloc[0]) if len(sample) > 0 else ""
            desc = f"HTF feature ({field.replace('htf_', '')})"
            ws_dict.append([field_num, field, desc, sample_val])
        uncategorized = [c for c in uncategorized if c not in htf_cols]

    if uncategorized:
        row_num += 1
        ws_dict.append([None, "── Other ──", None, None])
        for col_idx in range(1, 5):
            cell = ws_dict.cell(row=row_num, column=col_idx)
            cell.font = cat_font
            cell.fill = cat_fill
        for field in uncategorized:
            row_num += 1
            field_num += 1
            sample = df[field].dropna()
            sample_val = str(sample.iloc[0]) if len(sample) > 0 else ""
            ws_dict.append([field_num, field, FIELD_DESCRIPTIONS.get(field, ""), sample_val])

    # Freeze top row
    ws_dict.freeze_panes = "A2"

    # ── Sheet 2: Touch Data ───────────────────────────────────────────

    ws_data = wb.create_sheet("Touch Data")

    # Write headers
    for col_idx, col_name in enumerate(df.columns, 1):
        cell = ws_data.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="left")

    # Write data rows
    for r_idx, row in enumerate(df.itertuples(index=False), 2):
        for c_idx, value in enumerate(row, 1):
            ws_data.cell(row=r_idx, column=c_idx, value=value)

    # Auto-filter and freeze
    ws_data.auto_filter.ref = ws_data.dimensions
    ws_data.freeze_panes = "A2"

    # Save
    wb.save(xlsx_path)
    print(f"  Saved -> {xlsx_path}")
    print(f"  Sheet 1: Data Dictionary ({field_num} fields)")
    print(f"  Sheet 2: Touch Data ({len(df)} rows)")


if __name__ == "__main__":
    csv_path = (
        sys.argv[1]
        if len(sys.argv) > 1
        else "supertrend_touch_output/supertrend_touches_ALL_5m.csv"
    )
    xlsx_path = sys.argv[2] if len(sys.argv) > 2 else csv_path.replace(".csv", ".xlsx")
    export_to_excel(csv_path, xlsx_path)
